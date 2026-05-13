#!/usr/bin/env python3
"""
Migrate AFS/TNT Vacation Tracker Excel → Supabase.
Requires: pip install openpyxl requests
Env: NEXT_PUBLIC_SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY
"""

import openpyxl
import requests
import os
from datetime import date

SUPABASE_URL = os.environ['NEXT_PUBLIC_SUPABASE_URL']
SERVICE_KEY  = os.environ['SUPABASE_SERVICE_ROLE_KEY']
YEAR = 2026

HEADERS = {
    'apikey':        SERVICE_KEY,
    'Authorization': f'Bearer {SERVICE_KEY}',
    'Content-Type':  'application/json',
}

LEAVE_CODES = {'L','L1','L2','S','S1','S2','P','C','T','W','B'}

MONTHS = [
    'January','February','March','April','May','June',
    'July','August','September','October','November','December'
]

FILES = [
    {
        'path':    r'C:\Users\nero_\OneDrive - afstrans.co\afstrans.co - AFS_2023\Admin\2. HR\8. Vacation\Vacation Tracker 2026.xlsx',
        'company': 'AFS',
    },
    {
        'path':    r'C:\Users\nero_\OneDrive - afstrans.co\afstrans.co - AFS_2023\TNT\Admin\2. HR\Vacation\TNT Vacation Traker 2026.xlsx',
        'company': 'TNT',
    },
]

SKIP = ['manager','team','department','leave','total','work from','vacation','employee name']


def api_get(table, qs=''):
    r = requests.get(f'{SUPABASE_URL}/rest/v1/{table}?{qs}', headers=HEADERS)
    r.raise_for_status()
    return r.json()


def api_post(table, data, upsert=False):
    h = dict(HEADERS)
    h['Prefer'] = 'resolution=merge-duplicates,return=representation' if upsert else 'return=representation'
    r = requests.post(f'{SUPABASE_URL}/rest/v1/{table}', headers=h, json=data)
    r.raise_for_status()
    return r.json()


def get_company_id(fragment):
    rows = api_get('companies', f'name=ilike.*{fragment}*&select=id,name')
    if not rows:
        raise ValueError(f'Company not found: {fragment}')
    print(f'  Company: {rows[0]["name"]}  id={rows[0]["id"]}')
    return rows[0]['id']


def is_employee(val):
    if not val or not isinstance(val, str):
        return False
    v = val.strip()
    if len(v) < 3:
        return False
    vl = v.lower()
    if any(k in vl for k in SKIP):
        return False
    parts = v.split()
    if len(parts) == 2 and parts[0].lower() == 'employee' and parts[1].isdigit():
        return False
    return True


def find_date_row(ws):
    for rn in range(1, 15):
        row = list(ws.iter_rows(min_row=rn, max_row=rn))[0]
        m = {c.column: int(c.value) for c in row
             if isinstance(c.value, (int, float)) and 1 <= int(c.value) <= 31}
        if len(m) >= 20:
            return rn, m
    return None, {}


def parse_month(ws, year, month_num):
    entries = []
    _, col_day = find_date_row(ws)
    if not col_day:
        return entries

    manager, team = None, None

    for row in ws.iter_rows(min_row=6):
        col_a = row[0].value if len(row) > 0 else None
        col_b = row[1].value if len(row) > 1 else None

        # Detect team / manager labels
        for cell in row:
            if not cell.value or not isinstance(cell.value, str):
                continue
            v, vl = cell.value.strip(), cell.value.strip().lower()
            if 'manager:' in vl:
                manager = v[vl.index('manager:') + 8:].strip()
            elif vl.startswith(('team ', 'department')):
                team = v

        # Employee row
        if is_employee(col_b):
            name = col_b.strip()
            for cell in row:
                if cell.column not in col_day or not cell.value:
                    continue
                code = str(cell.value).strip().upper()
                if code in LEAVE_CODES:
                    try:
                        entries.append({
                            'name': name, 'team': team, 'manager': manager,
                            'date': date(year, month_num, col_day[cell.column]).isoformat(),
                            'code': code,
                        })
                    except ValueError:
                        pass
    return entries


def migrate(info):
    print(f'\n{"="*50}\nMigrating {info["company"]}\n{"="*50}')
    company_id = get_company_id(info['company'])
    wb = openpyxl.load_workbook(info['path'], data_only=True)

    all_entries, emp_meta = [], {}
    for mi, mname in enumerate(MONTHS, 1):
        if mname not in wb.sheetnames:
            continue
        rows = parse_month(wb[mname], YEAR, mi)
        all_entries.extend(rows)
        for r in rows:
            if r['name'] not in emp_meta:
                emp_meta[r['name']] = {'team': r['team'], 'manager': r['manager']}
        print(f'  {mname}: {len(rows)} entries')

    print(f'\nEmployees: {list(emp_meta.keys())}')
    emp_ids = {}

    for name, meta in emp_meta.items():
        existing = api_get('employees', f'company_id=eq.{company_id}&name=eq.{requests.utils.quote(name)}&select=id')
        if existing:
            emp_ids[name] = existing[0]['id']
            print(f'  [skip] {name}')
        else:
            r = api_post('employees', {
                'company_id': company_id, 'name': name,
                'team': meta['team'], 'manager_name': meta['manager'],
                'vacation_allowance': 24, 'is_active': True,
            })
            emp_ids[name] = r[0]['id']
            print(f'  [new]  {name}')

    payload = [
        {'employee_id': emp_ids[e['name']], 'date': e['date'], 'leave_code': e['code']}
        for e in all_entries if e['name'] in emp_ids
    ]

    print(f'\nUpserting {len(payload)} leave entries...')
    for i in range(0, len(payload), 500):
        api_post('leave_entries', payload[i:i+500], upsert=True)
        print(f'  {min(i+500, len(payload))}/{len(payload)}')

    print(f'{info["company"]} done.')


if __name__ == '__main__':
    for f in FILES:
        migrate(f)
    print('\nAll migrations complete.')